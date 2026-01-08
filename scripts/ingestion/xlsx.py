import openpyxl
from scripts.ingestion.models import AbstractIngestor, UnsupportedFileError


class XlsxIngestor(AbstractIngestor):
    """
    Ingestor for XLSX files. Splits each sheet into row-grouped text chunks.
    """

    def ingest(self, filepath: str) -> list[tuple[str, dict]]:
        if not filepath.endswith(".xlsx"):
            raise UnsupportedFileError("File is not a .xlsx file.")

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            extracted_data = []
            group_size = 50  # rows per chunk

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                current_chunk_lines = []
                chunk_start_row = 1
                row_count = 0

                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if not row:
                        continue

                    line = "\t".join(str(cell) for cell in row if cell is not None)
                    if not line.strip():
                        continue

                    current_chunk_lines.append(line)
                    row_count += 1

                    if row_count >= group_size:
                        chunk_text = "\n".join(current_chunk_lines)
                        meta = {
                            "doc_type": "xlsx",
                            "sheet_name": sheet_name,
                            "row_range": f"{chunk_start_row}-{i}",
                            "source_filepath": filepath,
                            "type": "sheet_content",
                        }
                        extracted_data.append((chunk_text, meta))

                        # Reset for next group
                        current_chunk_lines = []
                        chunk_start_row = i + 1
                        row_count = 0

                # flush remainder
                if current_chunk_lines:
                    chunk_text = "\n".join(current_chunk_lines)
                    meta = {
                        "doc_type": "xlsx",
                        "sheet_name": sheet_name,
                        "row_range": f"{chunk_start_row}-{i}",
                        "source_filepath": filepath,
                        "type": "sheet_content",
                    }
                    extracted_data.append((chunk_text, meta))

            return extracted_data

        except Exception as e:
            raise UnsupportedFileError(f"Failed to process XLSX file: {filepath}: {e}")
